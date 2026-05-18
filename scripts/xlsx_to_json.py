#!/usr/bin/env python3
"""Convert the Excel tracker into JSON files for the dashboard.

Run locally before pushing, or let the GitHub Action handle it.
"""
import json, sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, date

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'data' / 'Daesh_Claims_Tracker.xlsx'
OUT_CLAIMS = ROOT / 'data' / 'claims.json'
OUT_REF = ROOT / 'data' / 'reference.json'

def iso(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v)

def main():
    if not XLSX.exists():
        print(f'Missing: {XLSX}', file=sys.stderr); sys.exit(1)

    wb = load_workbook(XLSX, data_only=True)
    log = wb['Claims Log']
    ref = wb['Reference']

    headers = [c.value for c in log[1]]
    et_cols = [i for i, h in enumerate(headers) if h and str(h).startswith('Event Type')]
    tt_cols = [i for i, h in enumerate(headers) if h and str(h).startswith('Target Type')]
    wt_cols = [i for i, h in enumerate(headers) if h and str(h).startswith('Weapon Type')]
    idx = {h: i for i, h in enumerate(headers) if h}

    claims = []
    for row in log.iter_rows(min_row=2, values_only=True):
        event_date = row[idx['Event Date']]
        claim_date = row[idx['Claim Date']]
        if event_date is None and claim_date is None:
            continue
        c_iso = iso(claim_date)
        e_iso = iso(event_date)
        retro = bool(e_iso and c_iso and e_iso[:7] != c_iso[:7])
        rec = {
            'event_date': e_iso,
            'claim_date': c_iso,
            'claim_month': c_iso[:7] if c_iso else None,
            'event_month': e_iso[:7] if e_iso else None,
            'retroactive': retro,
            'actor': row[idx['Actor']] or None,
            'country': row[idx['Country']] or None,
            'location': row[idx['Location Details']] or None,
            'event_types': [row[i] for i in et_cols if row[i]],
            'target_types': [row[i] for i in tt_cols if row[i]],
            'weapon_types': [row[i] for i in wt_cols if row[i]],
            'fatalities': row[idx['Fatalities']] if row[idx['Fatalities']] is not None else 0,
            'summary': row[idx['Event Summary']] or '',
            'other': row[idx['Other']] or None,
        }
        claims.append(rec)

    ref_headers = [c.value for c in ref[1]]
    ref_data = {h: [] for h in ref_headers if h}
    for row in ref.iter_rows(min_row=2, values_only=True):
        for i, h in enumerate(ref_headers):
            if h and i < len(row) and row[i]:
                ref_data[h].append(row[i])

    OUT_CLAIMS.write_text(json.dumps(claims, indent=2, ensure_ascii=False))
    OUT_REF.write_text(json.dumps(ref_data, indent=2, ensure_ascii=False))
    print(f'Wrote {len(claims)} claims -> {OUT_CLAIMS}')
    print(f'Wrote reference -> {OUT_REF}')

if __name__ == '__main__':
    main()
